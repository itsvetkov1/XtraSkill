from io import BytesIO
from typing import Dict, Any
import openpyxl
from fastapi import HTTPException

from .base import DocumentParser
from ..file_validator import validate_zip_bomb


class ExcelParser(DocumentParser):
    """Excel parser with data type preservation."""

    def parse(self, file_bytes: bytes) -> Dict[str, Any]:
        """
        Extract text from Excel file, preserving data types.

        Returns dict with:
            - text: Full extracted text with sheet headers
            - summary: First 5000 chars for AI context
            - metadata: Sheet names, row counts, column headers
        """
        try:
            wb = openpyxl.load_workbook(
                BytesIO(file_bytes),
                read_only=True,
                data_only=True,
                keep_links=False
            )

            all_text = []
            metadata = {
                "sheet_names": [],
                "row_counts": {},
                "column_headers": {},
                "total_sheets": 0,
                "total_rows": 0,
            }

            for sheet in wb.worksheets:
                sheet_name = sheet.title
                metadata["sheet_names"].append(sheet_name)

                # Add sheet header
                all_text.append(f"[Sheet: {sheet_name}]")

                rows_data = []
                column_headers = []

                for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                    # Convert cell values to strings, preserving data types
                    cell_values = [
                        str(cell.value) if cell.value is not None else ""
                        for cell in row
                    ]

                    # Skip completely empty rows
                    if all(val == "" for val in cell_values):
                        continue

                    # First non-empty row is column headers
                    if row_idx == 1 and any(cell_values):
                        column_headers = cell_values

                    # Join cells with tab separator
                    row_text = "\t".join(cell_values)
                    rows_data.append(row_text)

                # Join rows with newlines
                sheet_text = "\n".join(rows_data)
                all_text.append(sheet_text)

                # Update metadata
                row_count = len(rows_data)
                metadata["row_counts"][sheet_name] = row_count
                metadata["total_rows"] += row_count
                metadata["column_headers"][sheet_name] = column_headers

            metadata["total_sheets"] = len(wb.worksheets)

            # Combine all sheets
            full_text = "\n\n".join(all_text)

            return {
                "text": full_text,
                "summary": self.create_ai_summary(full_text),
                "metadata": metadata
            }

        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"Failed to parse Excel file: {str(e)}"
            )

    def validate_security(self, file_bytes: bytes) -> None:
        """
        Validate Excel file security.

        Raises HTTPException for malformed files or zip bombs.
        """
        # Check for zip bomb (XLSX is a zip archive)
        validate_zip_bomb(file_bytes)

        # Try opening to catch malformed XLSX
        try:
            wb = openpyxl.load_workbook(
                BytesIO(file_bytes),
                read_only=True,
                data_only=True
            )
            # If we can open it, it's valid
            wb.close()
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"Malformed Excel file: {str(e)}"
            )
